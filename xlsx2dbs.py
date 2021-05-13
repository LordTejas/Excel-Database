from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


def create_workbook(name: str, columns: list, index_section=True) -> None:
    """
    Creates a workbook with given name and columns
    :param name: name of the xlsx file.
    :param columns: column names to be added to new workbook
    :param index_section: wether to leave space for index columns or not
    :return: None
    """

    # New Workbook
    wb = Workbook()

    # Grabbing First sheet
    ws = wb.active

    # length of columns
    num_columns = len(columns) + bool(index_section)

    df = pd.DataFrame(columns=columns)
    # print(df.to_string(index=False))

    for r in dataframe_to_rows(df, index=index_section, header=True):
        ws.append(r)

    for row in ws.iter_rows(min_row=1, max_col=num_columns, max_row=5, values_only=True):
        print(row)

    wb.save(name)


def update(path, data):
    pass


if __name__ == '__main__':

    column_data = ['id', 'name', 'age', 'score']
    n = len(column_data)

    create_workbook('sample.xlsx', column_data, False)
